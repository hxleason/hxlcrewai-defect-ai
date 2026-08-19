#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
知识库构建脚本
将 data/ 目录下的 Excel 源文件转换为 JSON 格式，供系统运行时加载。
需要安装 openpyxl:  pip install openpyxl
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ 缺少 openpyxl 库，请先安装：pip install openpyxl")
    sys.exit(1)

# 项目根目录（脚本位于 scripts/ 下，根目录即上一级）
ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data"

# Excel 文件名（根据你的实际文件名修改）
EXPERT_RULES_XLSX = "专家规则库.xlsx"
FAILURE_CASES_XLSX = "失效分析库.xlsx"

# 输出 JSON 文件名
EXPERT_RULES_JSON = "expert_rules.json"
FAILURE_CASES_JSON = "failure_cases.json"


def read_sheet_rows(file_path: Path, sheet_name: str = None) -> list[dict]:
    """
    读取 Excel 中第一个工作表（或指定工作表），返回一行一个字典的列表。
    表头作为字典的键。
    """
    wb = load_workbook(file_path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        # 跳过完全空的行
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rows.append(row)

    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else f"col_{i}" for i, cell in enumerate(rows[0])]
    data = []
    for raw in rows[1:]:
        # 补齐长度
        row_vals = list(raw) + [""] * (len(header) - len(raw))
        row_dict = {}
        for i, col in enumerate(header):
            val = row_vals[i]
            if val is None:
                val = ""
            row_dict[col] = str(val).strip()
        # 如果整行都是空字符串（可能因为空行但未被完全跳过），跳过
        if all(v == "" for v in row_dict.values()):
            continue
        data.append(row_dict)
    return data


def parse_int(value, default=1):
    """解析整数，失败则返回默认值"""
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def build_expert_rules() -> list[dict]:
    """读取专家规则库 Excel，返回规则列表"""
    xlsx_path = DATA_DIR / EXPERT_RULES_XLSX
    if not xlsx_path.exists():
        print(f"❌ 未找到文件：{xlsx_path}")
        sys.exit(1)

    rows = read_sheet_rows(xlsx_path)
    rules = []
    for row in rows:
        rule = {
            "rule_id": row.get("规则ID", ""),
            "rule_class": row.get("规则分类", ""),
            "if_condition": row.get("IF条件", ""),
            "then_action": row.get("THEN动作", ""),
            "action_explanation": row.get("动作解释", ""),
            "remark": row.get("备注", "")  # 备注中通常包含“来源：...”
        }
        # 只保留 rule_id 非空的条目
        if rule["rule_id"]:
            rules.append(rule)

    return rules


def build_failure_cases() -> list[dict]:
    """读取失效分析库 Excel，返回案例列表"""
    xlsx_path = DATA_DIR / FAILURE_CASES_XLSX
    if not xlsx_path.exists():
        print(f"❌ 未找到文件：{xlsx_path}")
        sys.exit(1)

    rows = read_sheet_rows(xlsx_path)
    cases = []
    for row in rows:
        case = {
            "case_id": row.get("情景ID", ""),
            "equipment_category": row.get("设备大类", ""),
            "equipment_subcategory": row.get("设备小类", ""),
            "failure_mode": row.get("失效模式", ""),
            "failure_phenomenon": row.get("失效现象", ""),
            "direct_cause": row.get("直接原因", ""),
            "root_cause": row.get("根本原因", ""),
            "failure_consequence": row.get("失效后果", ""),
            "risk_level": row.get("风险等级", ""),
            "corrective_measures": row.get("整改措施", ""),
            "severity": parse_int(row.get("严重度S", ""), 1),
            "occurrence": parse_int(row.get("发生度O", ""), 1),
            "detection": parse_int(row.get("探测度D", ""), 1),
            "sod_basis": row.get("S/O/D依据", ""),
            "source_type": row.get("来源类型", ""),
            "reference": row.get("参考文献", ""),
            "note": row.get("备注", "")  # 例如 "核心训练样本"
        }
        if case["case_id"]:
            cases.append(case)

    return cases


def main():
    print("🔄 开始构建知识库...\n")

    # 数据目录检查
    if not DATA_DIR.exists():
        print(f"❌ 数据目录不存在：{DATA_DIR}")
        sys.exit(1)

    # 生成专家规则 JSON
    rules = build_expert_rules()
    rules_json_path = DATA_DIR / EXPERT_RULES_JSON
    with open(rules_json_path, "w", encoding="utf-8") as f:
        json.dump({"rules": rules}, f, ensure_ascii=False, indent=2)
    print(f"✅ 专家规则库转换完成：{len(rules)} 条 → {rules_json_path}")

    # 生成失效案例 JSON
    cases = build_failure_cases()
    cases_json_path = DATA_DIR / FAILURE_CASES_JSON
    with open(cases_json_path, "w", encoding="utf-8") as f:
        json.dump({"cases": cases}, f, ensure_ascii=False, indent=2)
    print(f"✅ 失效案例库转换完成：{len(cases)} 条 → {cases_json_path}")

    print("\n🎉 知识库构建完成！")
    print("说明：")
    print("  - 规则 JSON 结构：{\"rules\": [ ... ]}")
    print("  - 案例 JSON 结构：{\"cases\": [ ... ]}")
    print("  - 系统运行时请读取这些 JSON 文件。")


if __name__ == "__main__":
    main()